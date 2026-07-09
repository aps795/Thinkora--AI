import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Create DataFrames
gov_data = pd.DataFrame([
    {"Position / Title": "President", "Current Office Holder": "Droupadi Murmu", "Notes": "15th President (Assumed office: July 2022)"},
    {"Position / Title": "Vice President", "Current Office Holder": "Jagdeep Dhankhar", "Notes": "14th Vice President"},
    {"Position / Title": "Prime Minister", "Current Office Holder": "Narendra Modi", "Notes": "Serving 3rd consecutive term; longest-serving elected PM as of June 2026"},
    {"Position / Title": "Chief Justice of India (CJI)", "Current Office Holder": "Justice Surya Kant", "Notes": "53rd CJI (Took oath: November 24, 2025)"},
    {"Position / Title": "Lok Sabha Speaker", "Current Office Holder": "Om Birla", "Notes": "18th Lok Sabha"},
    {"Position / Title": "Chief Election Commissioner", "Current Office Holder": "Rajiv Kumar", "Notes": "25th CEC"},
    {"Position / Title": "RBI Governor", "Current Office Holder": "Shaktikanta Das", "Notes": "25th Governor"}
])

econ_data = pd.DataFrame([
    {"Metric": "Nominal GDP", "2026 Value / Rank": "~$4.15 Trillion", "Context / Source": "6th largest global economy"},
    {"Metric": "GDP (PPP)", "2026 Value / Rank": "~$18.90 Trillion", "Context / Source": "3rd largest global economy"},
    {"Metric": "GDP Growth Rate", "2026 Value / Rank": "6.4% – 6.6%", "Context / Source": "IMF & RBI FY26 projections; fastest-growing major economy"},
    {"Metric": "Financial Year", "2026 Value / Rank": "April 1 to March 31", "Context / Source": "Standard Indian fiscal calendar"},
    {"Metric": "Currency", "2026 Value / Rank": "Indian Rupee (INR)", "Context / Source": "Sign: ₹"}
])

geo_data = pd.DataFrame([
    {"Attribute": "Capital", "Data Point": "New Delhi"},
    {"Attribute": "Total States", "Data Point": "28"},
    {"Attribute": "Total Union Territories", "Data Point": "8"},
    {"Attribute": "Total Area", "Data Point": "3.287 million sq. km (7th largest globally)"},
    {"Attribute": "Population", "Data Point": "~1.43+ Billion (Most populous globally)"},
    {"Attribute": "Standard Time", "Data Point": "IST (UTC +05:30)"},
    {"Attribute": "Major River Systems", "Data Point": "Ganga, Brahmaputra, Indus, Godavari, Krishna, Narmada"}
])

nat_data = pd.DataFrame([
    {"Category": "Flag", "National Symbol": "Tricolour: Saffron, White, Green with Ashoka Chakra"},
    {"Category": "Emblem", "National Symbol": "Lion Capital of Ashoka"},
    {"Category": "Anthem", "National Symbol": "Jana Gana Mana (composed by Rabindranath Tagore)"},
    {"Category": "Song", "National Symbol": "Vande Mataram (composed by Bankim Chandra Chatterjee)"},
    {"Category": "Animal", "National Symbol": "Bengal Tiger (Panthera tigris tigris)"},
    {"Category": "Bird", "National Symbol": "Indian Peafowl (Pavo cristatus)"},
    {"Category": "Flower", "National Symbol": "Lotus (Nelumbo nucifera)"},
    {"Category": "Tree", "National Symbol": "Banyan (Ficus benghalensis)"},
    {"Category": "River", "National Symbol": "Ganga"}
])

# Save to Excel via Pandas first
file_path = "India_GS_GK_Database_2026.xlsx"
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    gov_data.to_excel(writer, sheet_name='Governance', index=False)
    econ_data.to_excel(writer, sheet_name='Economic_Indicators', index=False)
    geo_data.to_excel(writer, sheet_name='Geo_Demographics', index=False)
    nat_data.to_excel(writer, sheet_name='National_Identity', index=False)

# Load the workbook to apply professional styling
wb = openpyxl.load_workbook(file_path)

# Styling configurations
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark muted blue
header_font = Font(color="FFFFFF", bold=True)
thin_border = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)
zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Freeze the top row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Style header and rows
    for row_idx, row in enumerate(ws.iter_rows()):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            if row_idx == 0:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif row_idx % 2 == 0:
                cell.fill = zebra_fill

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass

        adjusted_width = min((max_length + 2), 45)
        adjusted_width = max(adjusted_width, 15)
        ws.column_dimensions[column].width = adjusted_width

# Save the styled workbook
wb.save(file_path)
print(f"File generated: {file_path}")
